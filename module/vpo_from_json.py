import os
import commentjson as json
import zipfile
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from typing import Union, List, Dict, Optional
import pandas as pd
from copy import deepcopy
import re

class ExcelProcessor:
    def __init__(self, config_path: str|Path, template_path: Optional[str|Path] = None, output_dir: Optional[str|Path] = None) -> None:
        """
        Инициализация процессора с использованием xlwings
        :param config_path: Путь к файлу конфигурации
        :param template_path: Путь к шаблону (опционально)
        :param output_dir: Папка для сохранения результатов (опционально)
        """
        # Загрузка конфигурации
        with open(config_path, 'r', encoding='utf-8') as f:
            self.config = json.load(f)
        
        self.common_init = self.config["common"]
        self.common = self.config["common"]
        self.years_data = {k: v for k, v in self.config.items() if k != "common"}
        
        # Установка путей
        self.template_path = template_path or os.path.join(
            os.path.dirname(__file__), "templates", "template.xlsx"
        )
        self.first_col_after_template = None
        self.output_dir = output_dir or os.path.join(
            os.path.dirname(os.path.dirname(__file__)), "created_files"
        )
        
        # Создание папки для результатов
        os.makedirs(self.output_dir, exist_ok=True)

    def process_years(self, input_dir: str|Path, selected_years: Union[List[str], str] = "all", include_optional: bool = False) -> None:
        """
        Основной метод обработки данных
        :param input_dir: Папка с исходными данными
        :param selected_years: Список годов или "all"
        :param include_optional: Включать optional-файлы
        """

        years_to_process = self._get_years_to_process(selected_years)

        try:
            for year in years_to_process:
                year_dir = os.path.join(input_dir, "VPO_1_"+str(year))
                print(f"Обработка года: {year}, название папки/архива: {year_dir}")

                files = self._find_files(year_dir, include_optional)
                print(f"Найдено файлов для обработки: {len(files)}")

                output_path = self._get_output_path(year)

                # 1. Загружаем шаблон
                template_wb = self._load_template_workbook()

                # 2. Создаём новую книгу с листами из шаблона
                new_wb = self._create_new_workbook_with_template_sheets(template_wb)
                
                # 3. Корректируем common под особенности года
                self.common = self._deep_merge_dicts(self.common_init, self.years_data[year])

                for file_path in files:
                    # 4. Обрабатываем данные и заполняем листы
                    data_columns = self._process_data_for_year(new_wb, file_path, year)
                
                # 5. Создаём итоговый лист
                for summary_sheet in self.common['summary_sheet']:
                    self._create_summary_sheet(new_wb, summary_sheet)

                # 6. Удаляем лист-шаблон
                self._remove_sheet(new_wb)
                # 7. Сохраняем
                self._save_workbook(new_wb, output_path)
        except Exception as e:
            print(f"Критическая ошибка при обработке файла: {str(e)}")
            raise

    def _get_years_to_process(self, selected_years: Union[List[str], str]) -> List[str]:
        """Определяет какие годы нужно обработать"""
        if selected_years == "all":
            return list(self.years_data.keys())
        elif isinstance(selected_years, (list, tuple)):
            return [year for year in selected_years if year in self.years_data]
        return [selected_years] if selected_years in self.years_data else []

    def _find_files(self, path: str, include_optional: bool) -> List[str]:
        """Находит файлы для обработки с учетом фильтров"""
        zip_path = path+".zip"
        if zipfile.is_zipfile(zip_path):
            with zipfile.ZipFile(zip_path) as z:
                matched_files = []
                for zip_info in z.infolist():
                    if zip_info.is_dir():
                        continue
                    
                    if self._match_patterns(zip_info.filename, include_optional):
                        extracted_path = z.extract(zip_info, path=os.path.dirname(zip_path))
                        matched_files.append(extracted_path)
                
                return matched_files
        else:
            return [
                os.path.join(root, name)
                for root, _, files in os.walk(path)
                for name in files
                if self._match_patterns(name, include_optional)
            ]

    def _match_patterns(self, filename: str, include_optional: bool) -> bool:
        """Проверяет соответствие файла шаблонам имен"""
        required_patterns = [p.lower() for p in self.common["search_patterns"]["required"]]
        if not include_optional:
            required_ok = any(p in filename.lower()
                for p in required_patterns
            )
            return required_ok
        else:
            optional_patterns = [p.lower() for p in self.common["search_patterns"]["optional"]]
            optional_ok = any((p in filename.lower() and r in filename.lower())
                for p in optional_patterns
                for r in required_patterns
            )
            
            return optional_ok

    def _get_output_path(self, year: str) -> str:
        """Генерирует путь для выходного файла"""
        source_name = Path(self.template_path).stem.split("_")[:-1]
        output_name = f"{year}_{"_".join(source_name)}.xlsx"
        return os.path.join(self.output_dir, output_name)

    def _find_first_empty_column(self, sheet) -> str:
        """Находит первый пустой столбец в первой строке"""
        col = 1
        while sheet.cell(row=1, column=col).value is not None:
            col += 1
        return get_column_letter(col)
    
    def _copy_sheet_structure(self, src_sheet, dst_sheet) -> None:
        """Копирует данные и ширину столбцов"""
        # Копируем данные
        for row in src_sheet.iter_rows():
            for cell in row:
                dst_cell = dst_sheet[cell.coordinate]
                dst_cell.value = cell.value

        # Копируем ширину столбцов
        for col_idx in range(1, src_sheet.max_column + 1):
            col_letter = get_column_letter(col_idx)
            if src_sheet.column_dimensions[col_letter].width:
                dst_sheet.column_dimensions[col_letter].width = src_sheet.column_dimensions[col_letter].width


    def _fill_sheet(self, sheet, formulas: Dict, source_filename: str) -> str:
        """Заполняет лист данными и формулами"""
        
        # Находим первый пустой столбец
        start_col = self._find_first_empty_column(sheet)

        if not self.first_col_after_template:
            self.first_col_after_template = start_col

        # Заголовок — имя файла без расширения
        sheet[f"{start_col}1"].value = formulas["params"]["sheet_col_name"]
        
        # Извлекаем саму формулу
        if not isinstance(formulas, dict) or "formula" not in formulas:
            raise ValueError("formulas должен быть словарём с ключом 'formula'")

        base_formula = formulas["formula"]  # строка вида "=INDEX(..., A2, ...)"
        last_row = sheet.max_row
        
        sheet_exist = formulas["params"]["sheet_exist"] 

        if self.common["checking_list_existence"]:
            sheet_exist = self._sheet_exists(formulas["params"]["path_source_file"], formulas["params"]["sheet_name"])
        # Вставляем формулу во все строки, начиная со 2-й
        for row in range(2, last_row + 1):
            cell = sheet[f"{start_col}{row}"]
            if sheet_exist: 
                cell.value = base_formula.replace(f"={formulas["params"]["row_condition"]}", f"={formulas["params"]["row_condition"][:-1]}{row}")
            else:
                cell.value = "=NA()"
    
        return start_col
    
    def _prepare_summary_sheet(self, wb, summary_sheet):
        """Создаёт или берёт существующий итоговый лист и возвращает (sheet, start_col_idx, last_row)"""
        if summary_sheet["name"] in wb.sheetnames:
            sheet = wb[summary_sheet["name"]]
        else:
            sheet = wb.create_sheet(summary_sheet["name"])
            if wb.sheetnames[0] != summary_sheet["name"]:
                self._copy_sheet_structure(wb.worksheets[0], sheet)

        start_col = self._find_first_empty_column(sheet)
        return sheet, column_index_from_string(start_col), sheet.max_row
    
    def _collect_headers(self, wb):
        """Собирает все уникальные заголовки из выходных листов"""
        headers = []
        for sheet_name in self.common["output_sheets"]["name_aliases"].values():
            if sheet_name in wb.sheetnames:
                src = wb[sheet_name]
                col_idx = column_index_from_string(self.first_col_after_template)
                while True:
                    val = src.cell(1, col_idx).value
                    if not val:
                        break
                    if val not in headers:
                        headers.append(val)
                    col_idx += 1
        return headers

    def _get_sumif_parts_for_column(self, wb, header, cond_col, crit_col):
        """Возвращает список частей SUMIF для обычного заголовка"""
        parts = []
        for sheet_name in self.common["output_sheets"]["name_aliases"].values():
            if sheet_name in wb.sheetnames:
                src_sheet = wb[sheet_name]
                for col in range(column_index_from_string(self.first_col_after_template), src_sheet.max_column + 1):
                    if src_sheet.cell(1, col).value == header:
                        col_letter = get_column_letter(col)
                        parts.append(
                            f"SUMIF('{sheet_name}'!${cond_col}:${cond_col},${crit_col}{{row}},'{sheet_name}'!{col_letter}:{col_letter})"
                        )
                        break
        return parts
    
    def _get_sumif_parts_for_tag(self, wb, keywords, cond_col, crit_col):
        """Возвращает список частей SUMIF для группы ключевых слов"""
        parts_per_sheet = []
        for sheet_name in self.common["output_sheets"]["name_aliases"].values():
            if sheet_name not in wb.sheetnames:
                continue
            src_sheet = wb[sheet_name]
            matched_cols = [
                get_column_letter(col)
                for col in range(column_index_from_string(self.first_col_after_template), src_sheet.max_column + 1)
                if self._header_matches(str(src_sheet.cell(1, col).value or "").lower(), keywords)
            ]
            if matched_cols:
                if len(matched_cols) == 1:
                    parts_per_sheet.append(
                        f"SUMIF('{sheet_name}'!${cond_col}:${cond_col},${crit_col}{{row}},'{sheet_name}'!{matched_cols[0]}:{matched_cols[0]})"
                    )
                else:
                    inner = " + ".join(
                        f"SUMIF('{sheet_name}'!${cond_col}:${cond_col},${crit_col}{{row}},'{sheet_name}'!{c}:{c})"
                        for c in matched_cols
                    )
                    parts_per_sheet.append(f"({inner})")
        return parts_per_sheet

    def _create_summary_sheet(self, wb, summary_sheet: Dict) -> None:
        """
        Создает итоговый лист.
        Если group_mode=False -> классический постолбцовый режим.
        Если group_mode=True  -> группировка столбцов по ключевым словам.
        
        tag_groups в group_mode:
            [["москва"], ["москва", "заочная"]]
        Имя тега генерится автоматически через "_".
        """
        if summary_sheet["name"] == "None":
            return

        # Подготовка итогового листа
        sheet, start_col_idx, last_row = self._prepare_summary_sheet(wb, summary_sheet)

        start_col = self._find_first_empty_column(sheet)
        start_col_idx = column_index_from_string(start_col)
        last_row = sheet.max_row
        tag_groups = summary_sheet.get("tag_groups", None)
        cond_col = summary_sheet["range_col"]
        crit_col = summary_sheet["criteria_col"]

        # --- 1. Обычный режим (по каждому заголовку) ---
        if not tag_groups:
            all_headers = self._collect_headers(wb)
            # Заголовки
            for i, header in enumerate(all_headers, start=start_col_idx):
                sheet.cell(1, i, header)
                # Формулы SUMIF
            for col_idx, header in enumerate(all_headers, start=start_col_idx):
                for row in range(2, last_row + 1):
                    parts = [p.format(row=row)
                            for p in self._get_sumif_parts_for_column(wb, header, cond_col, crit_col)]
                    if parts:  # <-- этот if должен быть ВНУТРИ цикла по row!
                        sheet.cell(row, col_idx, f"={' + '.join(parts)}")

        # --- 2. Групповой режим ---
        else:
            # Преобразуем [["москва"], ["москва","заочная"]] → {"Москва": ["москва"], "Москва_заочная": ["москва","заочная"]}
            tag_dict = {"_".join(words).title(): words for words in tag_groups}
            # Заголовки — названия тегов
            for i, tag_name in enumerate(tag_dict.keys(), start=start_col_idx):
                sheet.cell(1, i, tag_name)
                # Формулы SUMIF по группам
            for col_idx, (tag_name, keywords) in enumerate(tag_dict.items(), start=start_col_idx):
                for row in range(2, last_row + 1):
                    parts = [p.format(row=row)
                            for p in self._get_sumif_parts_for_tag(wb, keywords, cond_col, crit_col)]
                    if parts:  # <-- тоже внутри цикла по row
                        sheet.cell(row, col_idx, f"={' + '.join(parts)}")

    def _header_matches(self, header: str, keywords: list[str]) -> bool:
        """
        Проверяет, что все ключевые слова встречаются в заголовке как отдельные слова (без учёта регистра).
        Дефис считается частью слова (например, 'очно-заочная').
        """
        # Разбиваем заголовок на слова: буквы/цифры/дефис — в одном токене
        words = re.findall(r"[A-Za-zА-Яа-яЁё0-9\-]+", header.lower())
        return all(kw.lower() in words for kw in keywords)

    def _load_template_workbook(self):
        """Загружает шаблон"""
        return load_workbook(self.template_path)

    def _create_new_workbook_with_template_sheets(self, template_wb):
        """Создаёт новую книгу и копирует листы из шаблона"""
        new_wb = Workbook()
        # Удаляем стандартный лист
        default_sheet = new_wb.active
        new_wb.remove(default_sheet)

        for sheet in template_wb.worksheets:
            new_sheet = new_wb.create_sheet(sheet.title)
            self._copy_sheet_structure(sheet, new_sheet)

        return new_wb

    def _nan_error_handle_(self, base_formula: str, 
                     handle_errors: bool = True,
                     error_replacement: str|int = 0) -> str:
        formula = base_formula
        
        if "error_handle" in self.common:
            error_replacement = self.common["error_handle"]
        # Обработка ошибок (поверх обработки NaN)
        if handle_errors:
            formula = f"IFERROR({formula}, {error_replacement})"
        
        return formula

    def _generate_formulas(self, source_file_path: str, sheet_name: str, params: Dict, level_code: int) -> Dict:
        """
        Генерирует формулу с фильтрацией по level_code (например, B=01)
        """
        # Получаем путь
        abs_source = os.path.abspath(source_file_path)
        abs_source_dir = os.path.dirname(abs_source)
        filename = os.path.basename(abs_source)
        
        params["path_source_file"] = f"{abs_source_dir}\\{filename}"
        params["sheet_name"] = sheet_name
        params["sheet_exist"] = True

        if sheet_name == "None":
            params["sheet_exist"] = False
            return {"formula": "=NA()",
                    "params": params
        }

        # Формируем ссылку на лист
        if abs_source_dir == ".":
            sheet_ref = f"[{filename}]{sheet_name}"
        else:
            sheet_ref = f"{abs_source_dir}\\[{filename}]{sheet_name}"
        full_ref = f"'{sheet_ref}'!"

        array = params["array"]  # например, $A$12:$W$467
        row_params = params["row_num"]
        col_params = params["columns_num"]

        # Извлекаем границы диапазона
        arr_parts = array.split(':')
        start_row = int(arr_parts[0].split('$')[-1])  # 12 из $A$12
        start_col_letter = arr_parts[0].split('$')[1]  # $A$12 → A
        end_col_letter = arr_parts[1].split('$')[1]    # $W$467 → W
        end_row = int(arr_parts[1].split('$')[-1])

        # Диапазоны
        key_col = row_params['looup_array']  # например, D
        key_col_range = f"{full_ref}${key_col}${start_row}:${key_col}${end_row}"  # $D$12:$D$467

        edu_col = self.common["rows_aliases"]["column"]  # B
        edu_col_range = f"{full_ref}${edu_col}${start_row}:${edu_col}${end_row}"  # $B$12:$B$467

        header_row_range = f"{full_ref}${start_col_letter}${start_row}:${end_col_letter}${start_row}"  # $A$12:$W$12

        # AGGREGATE для поиска строки с учётом level_code
        aggregate_part = (
        f"SUMPRODUCT(MAX("
        f"({key_col_range}={row_params['lookup_value']})*"
        f"({edu_col_range}={level_code})*"
        f"ROW({key_col_range})"
        f"))-{start_row-1}"
        )

        # Номер столбца
        col_match = f"MATCH({col_params['lookup_value']},{header_row_range},{col_params['match_type']})"

        # Финальная формула
        formula = f"INDEX({full_ref}{array},{aggregate_part},{col_match})"
        
        params["row_condition"] = row_params['lookup_value']
        
        return {
            "formula": f"={self._nan_error_handle_(formula)}",
            "params": params
        }

    def _process_data_for_year(self, new_wb, source_file_path: str, year: str) -> Dict:
        """
        Универсальный процессинг под новую структуру:
        - На какие листы разбивать ("to_list")
        - На какие колонки разбивать ("to_col"), если не указано — в качестве столбцов используется противоположная группа.
        - Названия листов подхватываются из name_aliases, если есть.
        """

        data_columns = {}
        source_filename = Path(source_file_path).name

        # Определяем группы по common
        to_list = self.common["output_sheets"]["to_list"]  # "rows_aliases" или "list_aliases"
        to_col = self.common["output_sheets"].get("to_col", "list_aliases" if to_list == "rows_aliases" else "rows_aliases")
        name_overrides = self.common["output_sheets"].get("name_aliases", {})

        # Листы в итоговом файле
        list_group = self.common[to_list]          # rows_aliases или list_aliases
        list_aliases = list_group["aliases"]

        # Столбцы в итоговом файле
        col_group = self.common[to_col]
        col_aliases = col_group["aliases"]
       
        for sheet_key, sheet_value in list_aliases.items():
            # Имя листа: если есть в name_aliases — берём его, иначе sheet_key
            sheet_name = name_overrides.get(sheet_key, sheet_key)
            if sheet_name not in new_wb.sheetnames:
                base_sheet = new_wb.worksheets[0]
                new_sheet = new_wb.create_sheet(sheet_name)
                self._copy_sheet_structure(base_sheet, new_sheet)
            else:
                new_sheet = new_wb[sheet_name]

            for col_key, col_value in col_aliases.items():
                # rows_aliases — это критерий для level_code (поддержка списка значений)
                # list_aliases — это имя листа-донора
                if to_list == "rows_aliases":
                    level_code = list_aliases[sheet_key]
                    sheet_source = col_value
                    list_key = col_key
                else:
                    level_code = col_value
                    sheet_source = sheet_value
                    list_key = sheet_key
                    
                if list_key not in self.common["funcs"]:
                    continue
                
                # Получаем параметры                                 
                params = self.common["funcs"][list_key]
                # Подпись столбца в итоговом файле
                params["sheet_col_name"] = (Path(source_filename).stem + "_" + col_key)

                # Если "actual_name" переопределяет sheet_source — берём его
                if isinstance(params.get("actual_name", None), str):
                    sheet_source = params["actual_name"]

                # level_code: список (например, [1,3]) — тогда прогоняем все, иначе — один
                if isinstance(level_code, list):
                    code_list = level_code
                else:
                    code_list = [level_code]
                
                for code in code_list:
                    
                    formulas = self._generate_formulas(
                        source_file_path=source_file_path,
                        sheet_name=sheet_source,
                        params=params,
                        level_code=code
                    )
                    
                    # Заполняем
                    try:
                        start_col = self._fill_sheet(new_sheet, formulas, source_filename)
                        data_columns.setdefault(sheet_name, []).append(start_col)
                    except Exception as e:
                        print(f"Ошибка заполнения листа {sheet_name}, столбца {col_key}: {str(e)}")

        return data_columns



    def _remove_sheet(self, wb, sheet_reference: Union[str, int] = "шаблон") -> None:
        """
        Удаляет лист из рабочей книги по названию или индексу
        
        :param wb: Рабочая книга (openpyxl.workbook.Workbook)
        :param sheet_reference: Название листа или его индекс (начиная с 0)
        """
        try:
            if isinstance(sheet_reference, str):
                # Удаление по названию
                if sheet_reference in wb.sheetnames:
                    sheet = wb[sheet_reference]
                    wb.remove(sheet)
            elif isinstance(sheet_reference, int):
                # Удаление по индексу
                if 0 <= sheet_reference < len(wb.sheetnames):
                    sheet = wb.worksheets[sheet_reference]
                    wb.remove(sheet)
        except Exception as e:
            print(f"Ошибка при удалении листа: {str(e)}")

    def _save_workbook(self, wb, output_path: str):
        """Сохраняет книгу"""
        try:
            wb.save(output_path)
            print(f"Файл успешно сохранён: {output_path}")
        except Exception as e:
            print(f"Ошибка сохранения: {str(e)}")
            raise

    def _get_data_type_by_sheet_name(self, sheet_name: str) -> str:
        """Возвращает логический тип по имени листа"""
        reverse_map = {v: k for k, v in self.common["list_aliases"]["aliases"].items()}
        return reverse_map.get(sheet_name, "unknown")

    def _delete_dogs(self, file_name: str):
        import subprocess

        ps_script = f"""
        $excel = New-Object -ComObject Excel.Application
        $excel.Visible = $false
        $wb = $excel.Workbooks.Open("{file_name}")

        foreach ($sheet in $wb.Worksheets) {{
            $sheet.UsedRange.Replace("@", "", [Type]::Missing, [Type]::Missing, $false)
        }}

        $wb.Save()
        $wb.Close()
        $excel.Quit()
        """

        subprocess.run(["powershell", "-Command", ps_script], shell=True)

    def _sheet_exists(self, file_path, sheet_name) -> bool:
        xl = pd.ExcelFile(file_path)
        if sheet_name in xl.sheet_names:
            return True
        print(f"Листа {sheet_name} нет в файле {file_path}")
        return False
    
    def _deep_merge_dicts(self, base: dict, override: dict) -> dict:
        """
        Рекурсивно объединяет два словаря.
        Значения из override перекрывают значения base.
        """
        result = deepcopy(base)

        for key, value in override.items():
            if (
                key in result
                and isinstance(result[key], dict)
                and isinstance(value, dict)
            ):
                # Если оба значения словари → рекурсивный merge
                result[key] = self._deep_merge_dicts(result[key], value)
            else:
                # Перезаписываем значение или добавляем новый ключ
                result[key] = deepcopy(value)

        return result